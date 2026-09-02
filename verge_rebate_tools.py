# -*- coding: utf-8 -*-
import datetime as _doc_dt
_DOC_YEAR = _doc_dt.date.today().year

f"""
Rebate Folder Tools
===================
Runs four operations in sequence (each can be toggled on/off):

  1. Store Rename          — fix store name typos in filenames (rules stored
                             in stores.json, editable via the UI)
  2. Add Suffix            — append a user-specified suffix (default: " 2026")
                             to Excel filenames that don't have it
  3. Delete Year Rows      — remove rows where column A contains a
                             user-specified year string (default: "2025")
  4. Convert Legacy Excel  — convert every .xls / .xlsm / .xlt / .xlsb in the
                             folder to modern .xlsx using real Excel (COM).
                             Originals can be kept or deleted.

Ship this file together with verge_icon.ico and Verge_Logo.png
in the same folder for the window/taskbar icon and header logo.

Developed by Abad Umair Channa | Copyright © {_DOC_YEAR}
"""

import os
import sys
import time
import json
import subprocess
import threading
import queue
import traceback
try:
    import tkinter as tk
except ImportError:
    import sys
    print("ERROR: tkinter is not available. Install Python from python.org (not Microsoft Store).")
    sys.exit(1)
from theme_manager import ThemeManager, apply_theme_to_window, get_copyright_year
from header_manager import FixedHeaderManager
from logo_handler import LogoHandler
from tkinter import ttk, filedialog, scrolledtext, messagebox
from pathlib import Path
from datetime import datetime, date
import base64
import tempfile

# ── Auto-install required packages BEFORE importing them ──────────────────
def _pip_install(pkg_name):
    # Never pip-install from inside a frozen EXE: all deps are bundled by
    # PyInstaller, and re-launching sys.executable would spawn another
    # instance of the app itself (window flooding).
    if getattr(sys, "frozen", False):
        return
    subprocess.run(
        [sys.executable, "-m", "pip", "install", pkg_name,
         "--quiet", "--disable-pip-version-check"],
        capture_output=True,
    )

# Check each dependency by its real import module name (pywin32 -> win32com,
# pillow -> PIL), not the pip package name, so this only runs pip when a
# package is genuinely missing.
for _mod, _pip in [
    ("openpyxl",  "openpyxl"),
    ("xlrd",      "xlrd==1.2.0"),
    ("xlwt",      "xlwt"),
    ("xlutils",   "xlutils"),
    ("win32com",  "pywin32"),
    ("PIL",       "pillow"),
]:
    try:
        __import__(_mod)
    except ImportError:
        _pip_install(_pip)

# ── Now import after ensuring packages are installed ──────────────────────
try:
    from openpyxl import load_workbook as _load_xlsx
except ImportError:
    _load_xlsx = None

try:
    import xlrd
    import xlwt
    from xlutils.copy import copy as xl_copy
except ImportError:
    xlrd = xlwt = xl_copy = None

try:
    import win32com.client as _win32com_client
except ImportError:
    _win32com_client = None

try:
    from PIL import Image as _PI, ImageTk as _PIT
except ImportError:
    _PI = _PIT = None


# ═══════════════════════════════════════════════════════════════════════════
# BRAND / WINDOW CONFIG — kept in sync with Verge_Inventory_Aging_Processor.pyw
# ═══════════════════════════════════════════════════════════════════════════
NAVY  = "#0B0E13"
EMBEDDED_LOGO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "embedded_logo_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "embedded_logo_b64.txt"), "r").read().strip()
EMBEDDED_ICON_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "embedded_icon_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "embedded_icon_b64.txt"), "r").read().strip()

RED   = "#2C5FE3"
WHITE = "#ffffff"
LIGHT = "#171A1F"
LOG_BG = "#10141B"
LOG_FG = "#C9D1DC"

ICON_ICO_NAME = "verge_icon.ico"
LOGO_PNG_NAME = "Verge_Logo.png"
COPYRIGHT_TEXT = f"Developed by Abad Umair Channa | Copyright © {date.today().year} | All rights reserved."
ICON_ICO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "icon_ico_b64.txt"), "r").read().strip() if not getattr(sys, "frozen", False) else open(os.path.join(getattr(sys, "_MEIPASS", "."), "assets", "icon_ico_b64.txt"), "r").read().strip()


def _script_dir() -> str:
    """Directory containing this .pyw (or .exe when frozen)."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def _resource_path(name: str) -> str:
    """Resolve a bundled resource (logo PNG) whether running from source or
    from a PyInstaller one-file EXE (extra files extract to _MEIPASS)."""
    if getattr(sys, "frozen", False):
        base = getattr(sys, "_MEIPASS", _script_dir())
        return os.path.join(base, name)
    return os.path.join(_script_dir(), name)




def _extract_embedded_icon(b64, filename):
    """Decode an embedded base64 icon to a temp file; return path or None."""
    try:
        if not b64:
            return None
        import base64 as _b64, tempfile, os
        target = os.path.join(tempfile.gettempdir(), filename)
        with open(target, "wb") as fh:
            fh.write(_b64.b64decode(b64))
        return target if os.path.isfile(target) else None
    except Exception:
        return None

def _set_window_icon(root):
    """Set taskbar + titlebar icon from embedded base64 ICO."""
    import base64, tempfile, atexit, os, sys

    # 1. Try sys._MEIPASS (PyInstaller onefile extraction dir)
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        ico_path = os.path.join(meipass, "verge_icon.ico")
        if os.path.exists(ico_path):
            try:
                root.iconbitmap(default=ico_path)
                root.after(200, lambda p=ico_path: root.iconbitmap(default=p))
                return
            except Exception:
                pass

    # 2. Try next to the exe/script
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    ico_path = os.path.join(base_dir, "verge_icon.ico")
    if os.path.exists(ico_path):
        try:
            root.iconbitmap(default=ico_path)
            root.after(200, lambda p=ico_path: root.iconbitmap(default=p))
            return
        except Exception:
            pass

    # 3. Decode EMBEDDED_ICON_B64 to %TEMP% (no spaces, always writable)
    try:
        data = base64.b64decode(EMBEDDED_ICON_B64.strip())
        tmp_dir = os.environ.get("TEMP", tempfile.gettempdir())
        ico_path = os.path.join(tmp_dir, "verge_app_icon.ico")
        with open(ico_path, "wb") as f:
            f.write(data)
        root.iconbitmap(default=ico_path)
        root.after(200, lambda p=ico_path: root.iconbitmap(default=p))
        return
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════
# STORE MANAGEMENT — stores.json (user-editable list of rename pairs)
# ═══════════════════════════════════════════════════════════════════════════
STORES_FILE = _resource_path("stores.json")

def _default_stores():
    """Sensible starting list — user can add/edit/delete via UI."""
    return [
        {"old": "Eliff",       "new": "E Iliff"},
        {"old": "FM 1960",     "new": "FM1960"},
        {"old": "Mckellips",   "new": "McKellips"},
        {"old": "N 19th",      "new": "N 19"},
        {"old": "N 35th",      "new": "N 35"},
        {"old": "N 51st",      "new": "N 51"},
        {"old": "New Thomas",  "new": "W Thomas"},
        {"old": "Mt View RD",  "new": "Mount View"},
        {"old": "4363 W Fuqua St", "new": "Fuqua"},
    ]

def load_stores():
    """Load store rename pairs from stores.json; fall back to defaults."""
    try:
        if os.path.exists(STORES_FILE):
            with open(STORES_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                return [(s.get("old", ""), s.get("new", "")) for s in data if s.get("old")]
    except Exception:
        pass
    # Save defaults so the file appears next to the exe for easy editing
    pairs = _default_stores()
    save_stores(pairs)
    return [(s["old"], s["new"]) for s in pairs]

def save_stores(pairs):
    """Save store rename pairs to stores.json."""
    try:
        data = [{"old": o, "new": n} for o, n in pairs]
        with open(STORES_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


def step1_store_rename(folder: Path, log, stores=None):
    log("\n── STEP 1: Store Rename ──────────────────────────────────")
    if stores is None:
        stores = load_stores()
    log(f"  Using {len(stores)} store rename rules.")
    renamed = skipped = 0
    for f in sorted(folder.iterdir()):
        if not f.is_file():
            continue
        new_name = f.name
        for old, new in stores:
            if old and old in new_name:
                new_name = new_name.replace(old, new)
        if new_name != f.name:
            dest = f.parent / new_name
            try:
                f.rename(dest)
                log(f"  Renamed: {f.name}  →  {new_name}")
                renamed += 1
            except Exception as e:
                log(f"  ERROR renaming {f.name}: {e}")
        else:
            skipped += 1
    log(f"  Done — {renamed} renamed, {skipped} unchanged.")
    return renamed


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2 — ADD YEAR SUFFIX (user-configurable year, default 2026)
# ═══════════════════════════════════════════════════════════════════════════
def step2_add_suffix(folder: Path, log, suffix="2026"):
    log(f"\n── STEP 2: Add ' {suffix}' Suffix ────────────────────────")
    renamed = skipped = 0
    for f in sorted(folder.iterdir()):
        if not f.is_file():
            continue
        if f.suffix.lower() not in (".xlsx", ".xls"):
            continue
        stem = f.stem
        if f" {suffix}" in stem:
            log(f"  Skipped (already has ' {suffix}'): {f.name}")
            skipped += 1
            continue
        new_name = f"{stem} {suffix}{f.suffix}"
        dest = f.parent / new_name
        try:
            f.rename(dest)
            log(f"  Renamed: {f.name}  →  {new_name}")
            renamed += 1
        except Exception as e:
            log(f"  ERROR renaming {f.name}: {e}")
    log(f"  Done — {renamed} renamed, {skipped} skipped.")
    return renamed


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3 — DELETE YEAR ROWS (user-configurable year, default 2025)
# ═══════════════════════════════════════════════════════════════════════════
def _delete_year_xlsx(path: Path, log, year="2025") -> str:
    """Delete rows where col A contains the year string from an .xlsx file."""
    if _load_xlsx is None:
        return "SKIP (openpyxl not installed)"
    wb = _load_xlsx(str(path))
    ws = wb.active
    rows_to_delete = [
        r for r in range(1, ws.max_row + 1)
        if year in str(ws.cell(r, 1).value or "")
    ]
    if not rows_to_delete:
        return f"SKIP (no {year} rows)"
    # Delete bottom-up so row indices stay valid
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
    wb.save(str(path))
    return f"OK — deleted {len(rows_to_delete)} row(s)"


def _delete_year_xls(path: Path, log, year="2025") -> str:
    """
    Delete rows where col A contains the year string from a legacy .xls file.
    Saves the result as .xlsx (modern format) so Excel / Power Query can open it.
    The original .xls file is removed after successful conversion.
    """
    if xlrd is None:
        return "SKIP (xlrd not installed)"
    if _load_xlsx is None:
        return "SKIP (openpyxl not installed)"

    # Read source with xlrd (only library that reads legacy .xls properly)
    rb = xlrd.open_workbook(str(path), formatting_info=False)
    rs = rb.sheet_by_index(0)

    keep_rows = [
        r for r in range(rs.nrows)
        if year not in str(rs.cell_value(r, 0))
    ]
    deleted = rs.nrows - len(keep_rows)
    if deleted == 0:
        return f"SKIP (no {year} rows)"

    # Write kept rows into a new .xlsx workbook via openpyxl
    import openpyxl
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = rs.name

    for new_r_idx, old_r in enumerate(keep_rows, start=1):
        for c in range(rs.ncols):
            cell_val = rs.cell_value(old_r, c)
            # Convert xlrd floats that look like integers back to int
            if isinstance(cell_val, float) and cell_val == int(cell_val):
                cell_val = int(cell_val)
            new_ws.cell(row=new_r_idx, column=c + 1, value=cell_val)

    # Save as .xlsx next to the original .xls file
    xlsx_path = path.with_suffix(".xlsx")
    new_wb.save(str(xlsx_path))

    # Remove the old .xls so only the clean .xlsx remains
    try:
        path.unlink()
    except Exception:
        pass

    return f"OK — deleted {deleted} row(s), saved as {xlsx_path.name}"


def step3_delete_year_rows(folder: Path, log, year="2025"):
    log(f"\n── STEP 3: Delete {year} Rows ────────────────────────────")
    modified = skipped = errors = 0
    # Process both .xls (convert to .xlsx) and .xlsx (edit in-place)
    xls_files = [f for f in sorted(folder.iterdir())
                 if f.is_file() and f.suffix.lower() in (".xls", ".xlsx")]
    if not xls_files:
        log("  No Excel files found.")
        return 0
    for f in xls_files:
        log(f"  Processing: {f.name}")
        max_retries = 3
        for attempt in range(1, max_retries + 1):
            try:
                ext = f.suffix.lower()
                if ext == ".xlsx":
                    result = _delete_year_xlsx(f, log, year)
                else:
                    result = _delete_year_xls(f, log, year)
                log(f"    {result}")
                if result.startswith("OK"):
                    modified += 1
                else:
                    skipped += 1
                break
            except Exception as e:
                log(f"    Attempt {attempt}/{max_retries} failed: {e}")
                if attempt < max_retries:
                    time.sleep(2)
                else:
                    log(f"    FINAL FAILURE — skipping file.")
                    errors += 1
    log(f"  Done — {modified} modified, {skipped} skipped, {errors} errors.")
    return modified


# ═══════════════════════════════════════════════════════════════════════════
# STEP 4 — CONVERT LEGACY EXCEL (.xls / .xlsm / .xlt / .xlsb → .xlsx)
# Merged in from verge_xls_to_xlsx.pyw so the rebate folder can be fully
# modernised in one pass. Uses real Microsoft Excel via COM so formatting,
# formulas and data are preserved exactly.
# ═══════════════════════════════════════════════════════════════════════════
XLSX_FILE_FORMAT = 51        # xlOpenXMLWorkbook
LEGACY_EXTS = (".xls", ".xlsm", ".xlt", ".xlsb", ".xlc")   # convert these → .xlsx
EXCEL_RESTART_EVERY = 60     # restart Excel periodically to avoid memory bloat

_CANCEL_CONVERT = threading.Event()


def _find_legacy_files(folder: Path, recurse: bool):
    out = []
    if recurse:
        for root, _, files in os.walk(folder):
            for f in files:
                out.append(Path(root) / f)
    else:
        out = [f for f in folder.iterdir() if f.is_file()]
    res = []
    for p in out:
        if p.name.startswith("~$"):
            continue                    # Excel lock files
        if p.suffix.lower() in LEGACY_EXTS:
            res.append(p)
    return sorted(res)


class _ExcelConverter:
    """Thin wrapper around the Excel COM instance that converts legacy
    Excel files to .xlsx. Recycles the Excel process periodically."""

    def __init__(self, log):
        self.log = log
        self.xl = None
        self._opened = 0

    def _start_excel(self):
        if _win32com_client is None:
            raise RuntimeError("pywin32 not installed — cannot drive Excel.")
        self.xl = _win32com_client.DispatchEx("Excel.Application")
        self.xl.Visible = False
        self.xl.DisplayAlerts = False
        try: self.xl.AutomationSecurity = 3   # block macros from prompting
        except Exception: pass
        try: self.xl.AskToUpdateLinks = False
        except Exception: pass

    def _stop_excel(self):
        if self.xl is not None:
            try: self.xl.Quit()
            except Exception: pass
        self.xl = None

    def _recycle_if_needed(self):
        self._opened += 1
        if self._opened % EXCEL_RESTART_EVERY == 0:
            self._stop_excel(); time.sleep(1); self._start_excel()

    def convert_one(self, path: Path, overwrite: bool, delete_original: bool) -> str:
        out = path.with_suffix(".xlsx")
        if out.exists() and not overwrite:
            return "skip"
        wb = None
        try:
            try:
                wb = self.xl.Workbooks.Open(os.fspath(path.resolve()),
                                            UpdateLinks=0, ReadOnly=True)
            except Exception:
                # corrupt/odd file → try Excel's repair-open
                wb = self.xl.Workbooks.Open(os.fspath(path.resolve()),
                                            UpdateLinks=0, ReadOnly=True,
                                            CorruptLoad=1)
            wb.SaveAs(os.fspath(out.resolve()), FileFormat=XLSX_FILE_FORMAT)
            wb.Close(False); wb = None
            self._recycle_if_needed()
            if delete_original:
                try: path.unlink()
                except Exception as e:
                    self.log(f"      (kept original — delete failed: {e})")
            return "ok"
        except Exception as e:
            if wb is not None:
                try: wb.Close(False)
                except Exception: pass
            # a bad file can wedge the instance — recycle it
            try: self._stop_excel(); self._start_excel()
            except Exception: pass
            return f"error: {e}"


def step4_convert_legacy_excel(folder: Path, log,
                               recurse: bool = False,
                               overwrite: bool = False,
                               delete_original: bool = False):
    """Convert every legacy .xls/.xlsm/.xlt/.xlsb in `folder` to .xlsx."""
    log("\n── STEP 4: Convert Legacy Excel → .xlsx ────────────────")
    if _win32com_client is None:
        log("  SKIP — pywin32 not installed. Install with: pip install pywin32")
        return 0

    _CANCEL_CONVERT.clear()
    files = _find_legacy_files(folder, recurse)
    if not files:
        log("  No legacy Excel files found in folder.")
        return 0

    log(f"  Found {len(files)} legacy Excel file(s). Starting Excel...")
    conv = _ExcelConverter(log)
    try:
        conv._start_excel()
    except Exception as e:
        log(f"  ERROR starting Excel: {e}")
        log("  Make sure Microsoft Excel is installed on this machine.")
        return 0

    ok = skip = err = 0
    try:
        for i, p in enumerate(files, 1):
            if _CANCEL_CONVERT.is_set():
                log("  ⏹ Cancelled by user.")
                break
            log(f"  [{i}/{len(files)}] {p.name}")
            r = conv.convert_one(p, overwrite, delete_original)
            if r == "ok":
                ok += 1
                log(f"      ✅ converted → {p.stem}.xlsx")
            elif r == "skip":
                skip += 1
                log(f"      ↷ skipped (.xlsx already exists)")
            else:
                err += 1
                log(f"      ❌ {r}")
    finally:
        conv._stop_excel()

    log(f"  Done — {ok} converted, {skip} skipped, {err} errors.")
    return ok


# ═══════════════════════════════════════════════════════════════════════════
# GUI  (styled to match Verge_Inventory_Aging_Processor.pyw)
# ═══════════════════════════════════════════════════════════════════════════
class App:
    def __init__(self, root):
        self.root = root
        self._q = queue.Queue()
        self._running = False
        self._logo_img = None

        root.title("Verge Desk Solutions - Rebate Tools")
        # Dynamic screen resolution support: size to 90% of the screen and
        # center it (DPI-aware), then stay a normal resizable top-level so
        # Windows Snap (50% left/right, corners, Win+arrow) keeps working.
        self._apply_dynamic_geometry()
        self.root.after(10, lambda: self.root.state("zoomed"))
        root.configure(bg=LIGHT)
        _set_window_icon(root)

        self.theme_manager = ThemeManager("Verge Rebate Folder Tools", app_name="verge-rebate-tools")
        self._styles()
        self._header()
        self._body()
        self._copyright_bar()
        apply_theme_to_window(self.root, self.theme_manager)
        self._poll()

    def _apply_dynamic_geometry(self) -> None:
        """Size the window to 90% of the screen and center it.

        Works on any laptop/monitor/PC (1080p, 1440p, 2K, 4K) and respects
        Windows DPI scaling (run after _enable_dpi_awareness()). The window
        stays resizable so Windows Snap gestures keep working — it centers
        on launch, then snaps normally to 50% left/right, corners or via
        Win+arrow shortcuts.
        """
        try:
            root = self.root
            root.update_idletasks()
            sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
            w = max(640, min(int(sw * 0.90), sw - 20))
            h = max(480, min(int(sh * 0.90), sh - 40))
            x = max(0, (sw - w) // 2)
            y = max(0, (sh - h) // 2)
            root.geometry(f"{w}x{h}+{x}+{y}")
            # minsize <= half the screen so 50% / corner snap is never blocked
            root.minsize(min(660, max(480, sw // 2)),
                         min(540, max(400, sh // 2)))
            root.resizable(True, True)
        except Exception:
            pass

    # ── styles ─────────────────────────────────────────────────────────────
    def _styles(self):
        s = ttk.Style(); s.theme_use("clam")
        s.configure("Run.TButton", background=RED, foreground=WHITE,
                    font=("Segoe UI", 11, "bold"), padding=(16, 9), borderwidth=0)
        s.map("Run.TButton",
              background=[("active", "#c01820"), ("disabled", "#aaa")])
        s.configure("Browse.TButton", background="#1E2228", foreground=WHITE,
                    font=("Segoe UI", 10), padding=(10, 6), borderwidth=0)
        s.map("Browse.TButton", background=[("active", "#171A1F")])
        s.configure("Cancel.TButton", background="#171A1F", foreground=WHITE,
                    font=("Segoe UI", 10), padding=(10, 6), borderwidth=0)
        s.map("Cancel.TButton", background=[("active", "#2A2C31")])
        s.configure("Accent.Horizontal.TProgressbar",
                    troughcolor="#1E2228", background=RED, borderwidth=0)

    # ── header (matches Aging Processor: NAVY 108px, logo left, title center) ──

    def _extract_embedded(self, b64, filename):
        """Decode an embedded base64 asset into a temp file; return path or None."""
        try:
            if not b64:
                return None
            import base64 as _b64, tempfile, os
            target = os.path.join(tempfile.gettempdir(), filename)
            with open(target, "wb") as fh:
                fh.write(_b64.b64decode(b64))
            return target if os.path.isfile(target) else None
        except Exception:
            return None


    def _lock_header_colors(self, widget, navy):
        """Recursively bind <Enter>/<Leave> on all header widgets to force navy."""
        try:
            widget.bind("<Enter>", lambda e, w=widget, c=navy: w.configure(bg=c) if not isinstance(w, type(None)) else None)
            widget.bind("<Leave>", lambda e, w=widget, c=navy: w.configure(bg=c) if not isinstance(w, type(None)) else None)
        except Exception:
            pass
        try:
            for child in widget.winfo_children():
                self._lock_header_colors(child, navy)
        except Exception:
            pass
    def _header(self):
        """Header using FixedHeaderManager with logo + theme toggle."""
        self.header_mgr = FixedHeaderManager(self.root, title="Verge Rebate Folder Tools")
        # Load the Verge logo into the header
        _logo_path = _resource_path(LOGO_PNG_NAME)
        if os.path.exists(_logo_path):
            self.header_mgr.set_logo(logo_path=_logo_path, text="Verge")
        # Add theme toggle button
        self.header_mgr.add_theme_toggle(self.theme_manager, callback=self._apply_theme)
        # FixedHeaderManager now tags ALL its own widgets with _tag="header"
        # in __init__/add_theme_toggle/add_copyright, so no manual tagging needed.


    def _apply_theme(self, colors=None):
        """Apply theme colors to all widgets EXCEPT header (header stays navy).

        Single source of truth: delegate to theme_manager.apply_theme_to_window(),
        which walks the tree, skips any widget with _tag in PROTECTED_TAGS,
        and handles Frame/Labelframe/Label/Button/Entry/Text/etc.
        """
        if colors is None:
            try:
                colors = self.theme_manager.get_colors()
            except Exception:
                return
        # theme_manager.apply_theme_to_window handles:
        #   - ttk.Style configuration (clam theme, TFrame/TLabel/TButton/etc.)
        #   - recursive _walk() that skips _tag-protected widgets (header)
        #   - Labelframe (was previously missed → panels stayed white)
        #   - Checkbutton/Radiobutton selectcolor
        self.theme_manager.apply_theme_to_window(self.root)
        # Refresh header toggle button text in case theme changed
        if hasattr(self.header_mgr, 'update_button_text'):
            self.header_mgr.update_button_text()


    def _body(self):
        body = tk.Frame(self.root, bg=LIGHT)
        body.pack(fill="both", expand=True, padx=24, pady=18)

        # Folder row
        row = tk.Frame(body, bg=LIGHT)
        row.pack(fill="x", pady=(0, 14))
        row.columnconfigure(0, weight=1)
        self.folder_var = tk.StringVar(value="")
        tk.Entry(row, textvariable=self.folder_var,
                 font=("Segoe UI", 9), relief="flat", bg="#1E2228", fg="#F5F7FA",
                 readonlybackground="#1E2228",
                 highlightbackground="#262B33", highlightthickness=1
                 ).grid(row=0, column=0, sticky="ew", ipady=5, padx=(0, 8))
        ttk.Button(row, text="Browse", style="Browse.TButton",
                   command=self._browse).grid(row=0, column=1)

        # Step checkboxes — 4 options now (Step 4 = Convert Legacy Excel)
        opts = tk.Frame(body, bg=LIGHT)
        opts.pack(fill="x", pady=(0, 10))
        self.do_step1 = tk.BooleanVar(value=True)
        self.do_step2 = tk.BooleanVar(value=True)
        self.do_step3 = tk.BooleanVar(value=True)
        self.do_step4 = tk.BooleanVar(value=True)
        for var, label in [
            (self.do_step1, "1. Store Rename"),
            (self.do_step2, "2. Add Suffix"),
            (self.do_step3, "3. Delete Year Rows"),
            (self.do_step4, "4. Convert Legacy Excel → .xlsx"),
        ]:
            tk.Checkbutton(opts, text=label, variable=var,
                           bg=LIGHT, fg="#F5F7FA", selectcolor=WHITE,
                           activebackground=LIGHT, activeforeground=NAVY,
                           font=("Segoe UI", 10)).pack(side="left", padx=(0, 16))

        # ── Year/Suffix entry fields (Step 2 suffix + Step 3 year) ───────────
        yr_row = tk.Frame(body, bg=LIGHT)
        yr_row.pack(fill="x", pady=(0, 10))
        tk.Label(yr_row, text="Step 2 — Suffix to add:",
                 bg=LIGHT, fg="#8A93A0", font=("Segoe UI", 9)).pack(side="left")
        self.suffix_var = tk.StringVar(value="2026")
        tk.Entry(yr_row, textvariable=self.suffix_var, width=8,
                 font=("Segoe UI", 9), relief="flat", bg="#1E2228", fg="#F5F7FA",
                 readonlybackground="#1E2228",
                 highlightbackground="#262B33", highlightthickness=1
                 ).pack(side="left", padx=(6, 20))
        tk.Label(yr_row, text="Step 3 — Year to delete:",
                 bg=LIGHT, fg="#8A93A0", font=("Segoe UI", 9)).pack(side="left")
        self.year_var = tk.StringVar(value="2025")
        tk.Entry(yr_row, textvariable=self.year_var, width=8,
                 font=("Segoe UI", 9), relief="flat", bg="#1E2228", fg="#F5F7FA",
                 readonlybackground="#1E2228",
                 highlightbackground="#262B33", highlightthickness=1
                 ).pack(side="left", padx=(6, 0))

        # ── Store management section (Step 1 rename rules) ───────────────────
        store_frame = tk.LabelFrame(body, text="Store Rename Rules (Step 1)",
                                    bg=LIGHT, fg="#F5F7FA", font=("Segoe UI", 9, "bold"),
                                    relief="flat", highlightbackground="#262B33",
                                    highlightthickness=1, padx=8, pady=6)
        store_frame.pack(fill="x", pady=(0, 10))

        # Treeview showing old → new pairs
        tree_holder = tk.Frame(store_frame, bg=LIGHT)
        tree_holder.pack(fill="x", pady=(0, 6))
        self.store_tree = ttk.Treeview(tree_holder, columns=("old", "new"),
                                       show="headings", height=5,
                                       selectmode="browse")
        self.store_tree.heading("old", text="Find (in filename)")
        self.store_tree.heading("new", text="Replace with")
        self.store_tree.column("old", width=200, anchor="w")
        self.store_tree.column("new", width=200, anchor="w")
        self.store_tree.pack(side="left", fill="x", expand=True)
        sb = ttk.Scrollbar(tree_holder, orient="vertical",
                           command=self.store_tree.yview)
        self.store_tree.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")

        # Add / Edit / Delete row
        edit_row = tk.Frame(store_frame, bg=LIGHT)
        edit_row.pack(fill="x", pady=(0, 6))
        tk.Label(edit_row, text="Find:", bg=LIGHT, fg="#8A93A0",
                 font=("Segoe UI", 9)).pack(side="left")
        self.new_old_var = tk.StringVar()
        tk.Entry(edit_row, textvariable=self.new_old_var, width=22,
                 font=("Segoe UI", 9), relief="flat", bg="#1E2228", fg="#F5F7FA",
                 highlightbackground="#262B33", highlightthickness=1
                 ).pack(side="left", padx=(4, 12))
        tk.Label(edit_row, text="Replace:", bg=LIGHT, fg="#8A93A0",
                 font=("Segoe UI", 9)).pack(side="left")
        self.new_new_var = tk.StringVar()
        tk.Entry(edit_row, textvariable=self.new_new_var, width=22,
                 font=("Segoe UI", 9), relief="flat", bg="#1E2228", fg="#F5F7FA",
                 highlightbackground="#262B33", highlightthickness=1
                 ).pack(side="left", padx=(4, 12))
        ttk.Button(edit_row, text="+ Add", style="Browse.TButton",
                   command=self._add_store).pack(side="left", padx=(0, 4))
        ttk.Button(edit_row, text="✎ Edit", style="Browse.TButton",
                   command=self._edit_store).pack(side="left", padx=(0, 4))
        ttk.Button(edit_row, text="✕ Delete", style="Browse.TButton",
                   command=self._delete_store).pack(side="left")

        # Load initial store list
        self._reload_store_tree()

        # Step-4 sub-options (only relevant when Step 4 is ticked)
        sub = tk.Frame(body, bg=LIGHT)
        sub.pack(fill="x", pady=(0, 10))
        self.recurse_var = tk.BooleanVar(value=True)
        self.overwrite_var = tk.BooleanVar(value=True)
        self.delete_original_var = tk.BooleanVar(value=True)
        for txt, var in [
            ("Include subfolders (Step 4)", self.recurse_var),
            ("Overwrite existing .xlsx (Step 4)", self.overwrite_var),
            ("Delete original after converting (Step 4)", self.delete_original_var),
        ]:
            tk.Checkbutton(sub, text=txt, variable=var,
                           bg=LIGHT, fg="#8A93A0", selectcolor=WHITE,
                           activebackground=LIGHT, activeforeground=NAVY,
                           font=("Segoe UI", 9)).pack(side="left", padx=(0, 14))

        # Run + Cancel buttons
        act = tk.Frame(body, bg=LIGHT)
        act.pack(fill="x", pady=(0, 10))
        self.run_btn = ttk.Button(act, text="▶  Run Selected Steps",
                                  style="Run.TButton", command=self._start)
        self.run_btn.pack(side="left")
        self.cancel_btn = ttk.Button(act, text="⏹  Cancel Step 4",
                                     style="Cancel.TButton",
                                     command=lambda: _CANCEL_CONVERT.set(),
                                     state="disabled")
        self.cancel_btn.pack(side="left", padx=8)
        self.status_var = tk.StringVar(value="Ready.")
        tk.Label(act, textvariable=self.status_var, bg=LIGHT, fg="#F5F7FA",
                 font=("Segoe UI", 9)).pack(side="left", padx=12)

        # Progress bar
        self.progress = ttk.Progressbar(body, mode="indeterminate",
                                        style="Accent.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(0, 10))

        # Log
        tk.Label(body, text="Activity Log", font=("Segoe UI", 9, "bold"),
                 fg="#F5F7FA", bg=LIGHT).pack(anchor="w")
        self.log_box = scrolledtext.ScrolledText(
            body, height=12, font=("Consolas", 8),
            bg=LOG_BG, fg=LOG_FG, relief="flat", state="disabled", wrap="word"
        )
        self.log_box.pack(fill="both", expand=True)

    def _copyright_bar(self):
        bar = tk.Frame(self.root, bg=NAVY, height=26)
        bar.pack(fill="x", side="bottom"); bar.pack_propagate(False)
        tk.Label(bar, text=COPYRIGHT_TEXT, bg=NAVY, fg="#8A93A0",
                 font=("Segoe UI", 8)).pack(pady=4)

    # ── helpers ────────────────────────────────────────────────────────────
    def _browse(self):
        d = filedialog.askdirectory(title="Select Rebate folder")
        if d:
            self.folder_var.set(d)

    # ── Store management helpers ──────────────────────────────────────────
    def _reload_store_tree(self):
        """Reload the store list from stores.json into the Treeview."""
        for item in self.store_tree.get_children():
            self.store_tree.delete(item)
        for old, new in load_stores():
            self.store_tree.insert("", "end", values=(old, new))

    def _add_store(self):
        """Add a new store rename pair from the entry fields."""
        old = self.new_old_var.get().strip()
        new = self.new_new_var.get().strip()
        if not old:
            messagebox.showwarning("Missing", "Please enter text to find in filenames.")
            return
        pairs = load_stores()
        # Prevent duplicates
        for o, n in pairs:
            if o == old:
                messagebox.showwarning("Duplicate", f"'{old}' already exists in the list.")
                return
        pairs.append((old, new))
        save_stores(pairs)
        self._reload_store_tree()
        self.new_old_var.set("")
        self.new_new_var.set("")

    def _edit_store(self):
        """Edit the selected store rename pair using the entry fields."""
        sel = self.store_tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select a row to edit first.")
            return
        old = self.new_old_var.get().strip()
        new = self.new_new_var.get().strip()
        if not old:
            messagebox.showwarning("Missing", "Please enter text to find in filenames.")
            return
        item = sel[0]
        old_old = self.store_tree.item(item, "values")[0]
        pairs = load_stores()
        new_pairs = []
        for o, n in pairs:
            if o == old_old:
                new_pairs.append((old, new))
            else:
                new_pairs.append((o, n))
        save_stores(new_pairs)
        self._reload_store_tree()
        self.new_old_var.set("")
        self.new_new_var.set("")

    def _delete_store(self):
        """Delete the selected store rename pair."""
        sel = self.store_tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select a row to delete first.")
            return
        item = sel[0]
        old_to_delete = self.store_tree.item(item, "values")[0]
        if not messagebox.askyesno("Confirm", f"Delete rule for '{old_to_delete}'?"):
            return
        pairs = load_stores()
        new_pairs = [(o, n) for o, n in pairs if o != old_to_delete]
        save_stores(new_pairs)
        self._reload_store_tree()

    def _log(self, msg):
        self._q.put(msg)

    def _poll(self):
        try:
            while True:
                msg = self._q.get_nowait()
                self.log_box.config(state="normal")
                self.log_box.insert("end", msg + "\n")
                self.log_box.see("end")
                self.log_box.config(state="disabled")
        except queue.Empty:
            pass
        self.root.after(80, self._poll)

    def _start(self):
        if self._running:
            return
        folder = Path(self.folder_var.get().strip())
        if not folder.exists():
            messagebox.showerror("Folder not found", str(folder))
            return
        # If Step 4 is ticked but pywin32 isn't available, warn early.
        if self.do_step4.get() and _win32com_client is None:
            if not messagebox.askyesno(
                "Excel driver missing",
                "Step 4 needs pywin32 + Microsoft Excel, which doesn't appear to be installed.\n\n"
                "Run Step 4 anyway? (It will skip itself if Excel can't start.)"
            ):
                return
        self._running = True
        self.run_btn.config(state="disabled")
        self.cancel_btn.config(state="normal")
        self.log_box.config(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.config(state="disabled")
        self.progress.start(12)
        threading.Thread(target=self._worker, args=(folder,), daemon=True).start()

    def _worker(self, folder: Path):
        try:
            self.status_var.set("Running…")
            self._log("═" * 55)
            self._log(f"Folder: {folder}")
            self._log(f"Steps: "
                      f"{'1 ' if self.do_step1.get() else ''}"
                      f"{'2 ' if self.do_step2.get() else ''}"
                      f"{'3 ' if self.do_step3.get() else ''}"
                      f"{'4' if self.do_step4.get() else ''}".rstrip())
            # Read the user-supplied suffix/year values (default to 2026/2025 if blank)
            suffix = self.suffix_var.get().strip() or "2026"
            year = self.year_var.get().strip() or "2025"
            if self.do_step2.get():
                self._log(f"  Step 2 suffix: '{suffix}'")
            if self.do_step3.get():
                self._log(f"  Step 3 year to delete: '{year}'")
            self._log("═" * 55)

            if self.do_step1.get():
                step1_store_rename(folder, self._log, stores=load_stores())
            if self.do_step2.get():
                step2_add_suffix(folder, self._log, suffix=suffix)
            if self.do_step3.get():
                step3_delete_year_rows(folder, self._log, year=year)
            if self.do_step4.get():
                step4_convert_legacy_excel(
                    folder, self._log,
                    recurse=self.recurse_var.get(),
                    overwrite=self.overwrite_var.get(),
                    delete_original=self.delete_original_var.get(),
                )

            self._log("")
            self._log("✓ COMPLETE — all selected steps finished.")
            self.status_var.set("Done.")
        except Exception as e:
            self._log(f"\nCRITICAL ERROR: {e}")
            self._log(traceback.format_exc())
            self.status_var.set("Error — see log.")
        finally:
            self.root.after(0, self._done)

    def _done(self):
        self.progress.stop()
        self.run_btn.config(state="normal")
        self.cancel_btn.config(state="disabled")
        self._running = False


# ═══════════════════════════════════════════════════════════════════════════
def _enable_dpi_awareness() -> None:
    """Make Windows report physical pixels so winfo_screen* is accurate on
    high-DPI displays (1080p, 1440p, 2K, 4K, DPI-scaled laptops)."""
    if sys.platform != "win32":
        return
    try:
        import ctypes
        # Set AppUserModelID BEFORE any window is created
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("VergeDesk.App")
        except Exception:
            pass
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)  # system DPI aware
        except Exception:
            ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass


if __name__ == "__main__":
    _enable_dpi_awareness()
    try:
        root = tk.Tk()
        App(root)
        root.mainloop()
    except Exception:
        traceback.print_exc()
        try:
            from tkinter import messagebox as _mb
            _mb.showerror("Fatal Error", traceback.format_exc())
        except Exception:
            pass
